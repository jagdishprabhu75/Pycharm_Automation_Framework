import openpyxl

DATA_FILE = "C:\\Misc\\users.xlsx"
usr_list = []


def read_excel():
    workbook = openpyxl.load_workbook(DATA_FILE)
    sheet = workbook["Sheet1"]
    row_count = sheet.max_row
    # col_count = sheet.max_column

    for r in range(2, row_count):
        username = sheet.cell(r, 1).value
        password = sheet.cell(r, 2).value
        tuple_user = (username, password)
        usr_list.append(tuple_user)

    return usr_list
